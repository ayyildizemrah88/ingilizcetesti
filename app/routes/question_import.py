# -*- coding: utf-8 -*-
"""
Question Import Routes
Bulk import questions from Excel/CSV
"""
import os
import tempfile
from flask import Blueprint, request, jsonify, render_template, flash, redirect, url_for, session
from werkzeug.utils import secure_filename

from app.extensions import db

question_import_bp = Blueprint('question_import', __name__)


def superadmin_required(f):
    """Decorator to require superadmin role."""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('rol') != 'superadmin':
            flash('Bu işlem için SuperAdmin yetkisi gereklidir.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@question_import_bp.route('/admin/sorular/import', methods=['GET', 'POST'])
@superadmin_required
def import_questions():
    """
    Bulk import questions from Excel or CSV file.
    """
    if request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('Dosya seçilmedi', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('Dosya seçilmedi', 'danger')
            return redirect(request.url)
        
        if not allowed_file(file.filename):
            flash('Geçersiz dosya formatı. Sadece .xlsx, .xls ve .csv kabul edilir.', 'danger')
            return redirect(request.url)
        
        try:
            # Save file temporarily
            filename = secure_filename(file.filename)
            temp_dir = tempfile.mkdtemp()
            filepath = os.path.join(temp_dir, filename)
            file.save(filepath)
            
            # Import questions
            from app.utils.question_importer import QuestionImporter
            
            importer = QuestionImporter()
            ext = os.path.splitext(filename)[1].lower()
            
            if ext in ['.xlsx', '.xls']:
                questions, stats = importer.import_from_excel(filepath)
            else:
                questions, stats = importer.import_from_csv(filepath)
            
            # Save to database
            if questions:
                saved, save_errors = importer.save_questions_to_db(questions)
                stats['saved'] = saved
                if save_errors:
                    stats['errors'].extend(save_errors)
            else:
                stats['saved'] = 0
            
            # Clean up
            os.remove(filepath)
            os.rmdir(temp_dir)
            
            # Show results
            if stats['saved'] > 0:
                flash(f"✅ {stats['saved']} soru başarıyla eklendi!", 'success')
            
            if stats['skipped'] > 0:
                flash(f"⚠️ {stats['skipped']} soru atlandı (hatalı veri)", 'warning')
            
            if stats['errors']:
                for error in stats['errors'][:5]:  # Show first 5 errors
                    flash(f"❌ {error}", 'danger')
                if len(stats['errors']) > 5:
                    flash(f"... ve {len(stats['errors']) - 5} hata daha", 'danger')
            
            return redirect(url_for('question_import.import_questions'))
            
        except Exception as e:
            flash(f'İçe aktarma hatası: {str(e)}', 'danger')
            return redirect(request.url)
    
    # GET request - show form
    from app.utils.question_importer import QuestionImporter
    importer = QuestionImporter()
    template_info = importer.get_template()
    
    return render_template('question_import.html', template_info=template_info)


@question_import_bp.route('/api/questions/import', methods=['POST'])
@superadmin_required
def api_import_questions():
    """
    API endpoint for question import.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya yüklenmedi'}), 400
    
    file = request.files['file']
    
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'error': 'Geçersiz dosya'}), 400
    
    try:
        # Save temporarily
        filename = secure_filename(file.filename)
        temp_dir = tempfile.mkdtemp()
        filepath = os.path.join(temp_dir, filename)
        file.save(filepath)
        
        # Import
        from app.utils.question_importer import import_questions_from_file
        saved, stats = import_questions_from_file(filepath)
        
        # Clean up
        os.remove(filepath)
        os.rmdir(temp_dir)
        
        return jsonify({
            'success': True,
            'saved': saved,
            'imported': stats.get('imported', 0),
            'skipped': stats.get('skipped', 0),
            'errors': stats.get('errors', [])
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@question_import_bp.route('/api/questions/template')
def get_import_template():
    """
    Get template information for question import.
    """
    from app.utils.question_importer import QuestionImporter
    importer = QuestionImporter()
    return jsonify(importer.get_template())


@question_import_bp.route('/admin/sorular/template/download')
@superadmin_required
def download_template():
    """
    Download Excel template for question import.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        import tempfile
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sorular"
        
        # Headers
        headers = [
            'soru_metni', 'kategori', 'seviye', 
            'sik_a', 'sik_b', 'sik_c', 'sik_d', 
            'dogru_cevap', 'aciklama', 'tags'
        ]
        
        # Style
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Example rows
        examples = [
            {
                'soru_metni': 'She ___ to work every day.',
                'kategori': 'grammar',
                'seviye': 'A2',
                'sik_a': 'go',
                'sik_b': 'goes',
                'sik_c': 'going',
                'sik_d': 'gone',
                'dogru_cevap': 'B',
                'aciklama': 'Third person singular present tense requires -s ending',
                'tags': 'verbs, present simple'
            },
            {
                'soru_metni': 'The opposite of "happy" is:',
                'kategori': 'vocabulary',
                'seviye': 'A1',
                'sik_a': 'glad',
                'sik_b': 'joyful',
                'sik_c': 'sad',
                'sik_d': 'excited',
                'dogru_cevap': 'C',
                'aciklama': 'Antonyms - opposite meanings',
                'tags': 'adjectives, antonyms'
            },
            {
                'soru_metni': 'I have been living here ___ 2010.',
                'kategori': 'grammar',
                'seviye': 'B1',
                'sik_a': 'for',
                'sik_b': 'since',
                'sik_c': 'from',
                'sik_d': 'during',
                'dogru_cevap': 'B',
                'aciklama': 'Since is used with specific points in time',
                'tags': 'present perfect, prepositions'
            }
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=example.get(header, ''))
                cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        for col in ['D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 25
        
        # Add instructions sheet
        ws2 = wb.create_sheet(title="Talimatlar")
        instructions = [
            ["Soru İçe Aktarma Talimatları"],
            [""],
            ["Gerekli Alanlar:"],
            ["- soru_metni: Sorunun tam metni"],
            ["- kategori: grammar, vocabulary, reading, listening, writing, speaking"],
            ["- seviye: A1, A2, B1, B2, C1, C2"],
            [""],
            ["Opsiyonel Alanlar:"],
            ["- sik_a, sik_b, sik_c, sik_d: Çoktan seçmeli şıklar"],
            ["- dogru_cevap: A, B, C veya D"],
            ["- aciklama: Doğru cevabın açıklaması"],
            ["- tags: Virgülle ayrılmış etiketler"],
            [""],
            ["İpuçları:"],
            ["- Her satır bir soru"],
            ["- İlk satır başlık satırıdır, silmeyin"],
            ["- Örnek satırları silip kendi sorularınızı ekleyin"],
        ]
        
        for row_num, row in enumerate(instructions, 1):
            ws2.cell(row=row_num, column=1, value=row[0] if row else '')
        
        ws2.column_dimensions['A'].width = 60
        
        # Save to temp file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        from flask import send_file
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='soru_sablonu.xlsx'
        )
        
    except ImportError:
        return jsonify({'error': 'openpyxl yüklü değil'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
