# -*- coding: utf-8 -*-
"""
Excel Export Utilities
Export reports and data to Excel format
"""
import os
import tempfile
from datetime import datetime
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """Export data to professionally formatted Excel files."""
    
    # Brand colors
    HEADER_COLOR = "667eea"
    HEADER_FONT_COLOR = "FFFFFF"
    ALT_ROW_COLOR = "f8f9fa"
    BORDER_COLOR = "dee2e6"
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup reusable styles."""
        self.header_fill = PatternFill(
            start_color=self.HEADER_COLOR,
            end_color=self.HEADER_COLOR,
            fill_type="solid"
        )
        self.header_font = Font(
            color=self.HEADER_FONT_COLOR,
            bold=True,
            size=11
        )
        self.alt_fill = PatternFill(
            start_color=self.ALT_ROW_COLOR,
            end_color=self.ALT_ROW_COLOR,
            fill_type="solid"
        )
        self.thin_border = Border(
            left=Side(style='thin', color=self.BORDER_COLOR),
            right=Side(style='thin', color=self.BORDER_COLOR),
            top=Side(style='thin', color=self.BORDER_COLOR),
            bottom=Side(style='thin', color=self.BORDER_COLOR)
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
    
    def _apply_header_style(self, row):
        """Apply header styling to a row."""
        for cell in row:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border
    
    def _apply_data_style(self, row, row_num):
        """Apply data row styling."""
        for cell in row:
            if row_num % 2 == 0:
                cell.fill = self.alt_fill
            cell.border = self.thin_border
            cell.alignment = self.left_align
    
    def _auto_width(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_candidates(self, candidates: List[Dict], filename: str = None) -> str:
        """
        Export candidates list to Excel.
        
        Args:
            candidates: List of candidate dictionaries
            filename: Optional output filename
        
        Returns:
            str: Path to generated Excel file
        """
        ws = self.ws
        ws.title = "Adaylar"
        
        # Headers
        headers = [
            'ID', 'Ad Soyad', 'E-posta', 'Telefon', 'Şirket',
            'Durum', 'Toplam Puan', 'CEFR Seviye',
            'Grammar', 'Vocabulary', 'Reading', 'Listening', 'Writing', 'Speaking',
            'Sınav Başlangıç', 'Sınav Bitiş', 'Oluşturulma'
        ]
        ws.append(headers)
        self._apply_header_style(ws[1])
        
        # Data rows
        for i, c in enumerate(candidates, 2):
            row = [
                c.get('id'),
                c.get('ad_soyad'),
                c.get('email'),
                c.get('telefon'),
                c.get('sirket_adi'),
                c.get('durum'),
                c.get('toplam_puan'),
                c.get('cefr_seviye'),
                c.get('grammar_puan'),
                c.get('vocabulary_puan'),
                c.get('reading_puan'),
                c.get('listening_puan'),
                c.get('writing_puan'),
                c.get('speaking_puan'),
                c.get('sinav_baslama'),
                c.get('sinav_bitis'),
                c.get('created_at')
            ]
            ws.append(row)
            self._apply_data_style(ws[i], i)
        
        self._auto_width(ws)
        
        # Save
        if not filename:
            filename = f"adaylar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(tempfile.gettempdir(), filename)
        self.wb.save(filepath)
        
        return filepath
    
    def export_exam_results(self, results: List[Dict], filename: str = None) -> str:
        """Export exam results to Excel."""
        ws = self.ws
        ws.title = "Sınav Sonuçları"
        
        headers = [
            'Aday', 'E-posta', 'Şirket', 'Sınav Tarihi',
            'Toplam Puan', 'CEFR Seviye',
            'Grammar', 'Vocabulary', 'Reading', 'Listening', 'Writing', 'Speaking',
            'Süre (dk)', 'Durum'
        ]
        ws.append(headers)
        self._apply_header_style(ws[1])
        
        for i, r in enumerate(results, 2):
            row = [
                r.get('ad_soyad'),
                r.get('email'),
                r.get('sirket_adi'),
                r.get('sinav_tarihi'),
                r.get('toplam_puan'),
                r.get('cefr_seviye'),
                r.get('grammar_puan'),
                r.get('vocabulary_puan'),
                r.get('reading_puan'),
                r.get('listening_puan'),
                r.get('writing_puan'),
                r.get('speaking_puan'),
                r.get('sure_dakika'),
                r.get('durum')
            ]
            ws.append(row)
            self._apply_data_style(ws[i], i)
        
        self._auto_width(ws)
        
        if not filename:
            filename = f"sinav_sonuclari_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(tempfile.gettempdir(), filename)
        self.wb.save(filepath)
        
        return filepath
    
    def export_analytics(self, data: Dict, filename: str = None) -> str:
        """Export analytics data with multiple sheets."""
        # Overview sheet
        ws = self.ws
        ws.title = "Genel Bakış"
        
        overview = [
            ['Metrik', 'Değer'],
            ['Toplam Aday', data.get('total_candidates', 0)],
            ['Tamamlanan Sınavlar', data.get('completed_exams', 0)],
            ['Ortalama Puan', f"{data.get('average_score', 0):.1f}"],
            ['Başarı Oranı', f"{data.get('pass_rate', 0):.1f}%"],
            ['Bu Ay Sınavlar', data.get('exams_this_month', 0)],
            ['Aktif Şirketler', data.get('active_companies', 0)],
        ]
        
        for row in overview:
            ws.append(row)
        
        self._apply_header_style(ws[1])
        for i in range(2, len(overview) + 1):
            self._apply_data_style(ws[i], i)
        
        self._auto_width(ws)
        
        # CEFR Distribution sheet
        ws2 = self.wb.create_sheet("CEFR Dağılımı")
        cefr_data = data.get('cefr_distribution', {})
        
        ws2.append(['Seviye', 'Aday Sayısı', 'Yüzde'])
        self._apply_header_style(ws2[1])
        
        total = sum(cefr_data.values())
        for i, (level, count) in enumerate(cefr_data.items(), 2):
            pct = (count / total * 100) if total > 0 else 0
            ws2.append([level, count, f"{pct:.1f}%"])
            self._apply_data_style(ws2[i], i)
        
        self._auto_width(ws2)
        
        # Skills sheet
        ws3 = self.wb.create_sheet("Beceri Ortalamaları")
        skills = data.get('skill_averages', {})
        
        ws3.append(['Beceri', 'Ortalama Puan'])
        self._apply_header_style(ws3[1])
        
        for i, (skill, avg) in enumerate(skills.items(), 2):
            ws3.append([skill.title(), f"{avg:.1f}"])
            self._apply_data_style(ws3[i], i)
        
        self._auto_width(ws3)
        
        if not filename:
            filename = f"analitik_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(tempfile.gettempdir(), filename)
        self.wb.save(filepath)
        
        return filepath
    
    def export_revenue_report(self, data: Dict, filename: str = None) -> str:
        """Export revenue report to Excel."""
        ws = self.ws
        ws.title = "Gelir Raporu"
        
        # Summary
        ws.append(['Gelir Özeti'])
        ws.append([])
        
        summary = [
            ['Toplam Gelir', f"₺{data.get('total_revenue', 0):,.2f}"],
            ['Bu Ay Gelir', f"₺{data.get('monthly_revenue', 0):,.2f}"],
            ['Toplam Sınav', data.get('total_exams', 0)],
            ['Sınav Başı Gelir', f"₺{data.get('revenue_per_exam', 0):,.2f}"],
        ]
        
        for row in summary:
            ws.append(row)
        
        ws.append([])
        ws.append(['Şirket Bazlı Gelir'])
        ws.append(['Şirket', 'Sınav Sayısı', 'Gelir'])
        
        companies = data.get('by_company', [])
        for c in companies:
            ws.append([c['name'], c['exams'], f"₺{c['revenue']:,.2f}"])
        
        self._auto_width(ws)
        
        if not filename:
            filename = f"gelir_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(tempfile.gettempdir(), filename)
        self.wb.save(filepath)
        
        return filepath


def export_to_excel(data_type: str, data: Any, filename: str = None) -> str:
    """
    Convenience function to export data to Excel.
    
    Args:
        data_type: Type of data ('candidates', 'results', 'analytics', 'revenue')
        data: Data to export
        filename: Optional output filename
    
    Returns:
        str: Path to generated Excel file
    """
    exporter = ExcelExporter()
    
    if data_type == 'candidates':
        return exporter.export_candidates(data, filename)
    elif data_type == 'results':
        return exporter.export_exam_results(data, filename)
    elif data_type == 'analytics':
        return exporter.export_analytics(data, filename)
    elif data_type == 'revenue':
        return exporter.export_revenue_report(data, filename)
    else:
        raise ValueError(f"Unknown data type: {data_type}")
