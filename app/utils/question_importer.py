# -*- coding: utf-8 -*-
"""
Bulk Question Importer
Import questions from Excel/CSV files
"""
import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Tuple, Optional

logger = logging.getLogger(__name__)


class QuestionImporter:
    """Import questions from Excel or CSV files."""
    
    REQUIRED_COLUMNS = ['soru_metni', 'kategori', 'seviye']
    OPTIONAL_COLUMNS = ['sik_a', 'sik_b', 'sik_c', 'sik_d', 'dogru_cevap', 'aciklama', 'tags']
    
    VALID_CATEGORIES = ['grammar', 'vocabulary', 'reading', 'listening', 'writing', 'speaking']
    VALID_LEVELS = ['A1', 'A2', 'B1', 'B2', 'C1', 'C2']
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        self.skipped_count = 0
    
    def validate_row(self, row: Dict, row_num: int) -> Tuple[bool, Dict]:
        """Validate a single row of question data."""
        errors = []
        cleaned = {}
        
        for col in self.REQUIRED_COLUMNS:
            value = row.get(col, '').strip() if row.get(col) else ''
            if not value:
                errors.append(f"Satır {row_num}: '{col}' alanı zorunludur")
            else:
                cleaned[col] = value
        
        kategori = cleaned.get('kategori', '').lower()
        if kategori and kategori not in self.VALID_CATEGORIES:
            errors.append(f"Satır {row_num}: Geçersiz kategori '{kategori}'")
        else:
            cleaned['kategori'] = kategori
        
        seviye = cleaned.get('seviye', '').upper()
        if seviye and seviye not in self.VALID_LEVELS:
            errors.append(f"Satır {row_num}: Geçersiz seviye '{seviye}'")
        else:
            cleaned['seviye'] = seviye
        
        for col in self.OPTIONAL_COLUMNS:
            value = row.get(col, '')
            if value:
                cleaned[col] = str(value).strip()
        
        if all(cleaned.get(f'sik_{x}') for x in ['a', 'b', 'c', 'd']):
            dogru = cleaned.get('dogru_cevap', '').upper()
            if dogru and dogru not in ['A', 'B', 'C', 'D']:
                errors.append(f"Satır {row_num}: Geçersiz doğru cevap '{dogru}'")
            cleaned['dogru_cevap'] = dogru
        
        if cleaned.get('tags'):
            try:
                if isinstance(cleaned['tags'], str):
                    if cleaned['tags'].startswith('['):
                        cleaned['tags'] = json.loads(cleaned['tags'])
                    else:
                        cleaned['tags'] = [t.strip() for t in cleaned['tags'].split(',')]
            except:
                cleaned['tags'] = [cleaned['tags']]
        
        if errors:
            self.errors.extend(errors)
            return False, {}
        
        return True, cleaned
    
    def import_from_excel(self, file_path: str) -> Tuple[List[Dict], Dict]:
        """Import questions from an Excel file."""
        try:
            import openpyxl
        except ImportError:
            self.errors.append("openpyxl kütüphanesi yüklü değil")
            return [], self._get_stats()
        
        if not os.path.exists(file_path):
            self.errors.append(f"Dosya bulunamadı: {file_path}")
            return [], self._get_stats()
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).lower().strip() if cell.value else '')
            
            missing_headers = [h for h in self.REQUIRED_COLUMNS if h not in headers]
            if missing_headers:
                self.errors.append(f"Eksik sütunlar: {', '.join(missing_headers)}")
                return [], self._get_stats()
            
            questions = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                
                is_valid, cleaned = self.validate_row(row_dict, row_num)
                
                if is_valid:
                    questions.append(cleaned)
                    self.imported_count += 1
                else:
                    self.skipped_count += 1
            
            wb.close()
            return questions, self._get_stats()
            
        except Exception as e:
            self.errors.append(f"Excel okuma hatası: {str(e)}")
            return [], self._get_stats()
    
    def import_from_csv(self, file_path: str, encoding='utf-8', delimiter=',') -> Tuple[List[Dict], Dict]:
        """Import questions from a CSV file."""
        import csv
        
        if not os.path.exists(file_path):
            self.errors.append(f"Dosya bulunamadı: {file_path}")
            return [], self._get_stats()
        
        try:
            questions = []
            
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                headers = [h.lower().strip() for h in reader.fieldnames]
                
                missing_headers = [h for h in self.REQUIRED_COLUMNS if h not in headers]
                if missing_headers:
                    self.errors.append(f"Eksik sütunlar: {', '.join(missing_headers)}")
                    return [], self._get_stats()
                
                for row_num, row in enumerate(reader, start=2):
                    row_dict = {k.lower().strip(): v for k, v in row.items()}
                    
                    is_valid, cleaned = self.validate_row(row_dict, row_num)
                    
                    if is_valid:
                        questions.append(cleaned)
                        self.imported_count += 1
                    else:
                        self.skipped_count += 1
            
            return questions, self._get_stats()
            
        except Exception as e:
            self.errors.append(f"CSV okuma hatası: {str(e)}")
            return [], self._get_stats()
    
    def save_questions_to_db(self, questions: List[Dict]) -> Tuple[int, List[str]]:
        """Save imported questions to database."""
        from app.extensions import db
        from app.models.question import Question
        
        saved = 0
        errors = []
        
        for q in questions:
            try:
                question = Question(
                    soru_metni=q['soru_metni'],
                    kategori=q['kategori'],
                    seviye=q['seviye'],
                    sik_a=q.get('sik_a'),
                    sik_b=q.get('sik_b'),
                    sik_c=q.get('sik_c'),
                    sik_d=q.get('sik_d'),
                    dogru_cevap=q.get('dogru_cevap'),
                    aciklama=q.get('aciklama'),
                    tags=json.dumps(q.get('tags', [])) if q.get('tags') else None,
                    aktif=True,
                    created_at=datetime.utcnow()
                )
                
                db.session.add(question)
                saved += 1
                
            except Exception as e:
                errors.append(f"Soru kaydedilemedi: {str(e)}")
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors.append(f"Veritabanı hatası: {str(e)}")
            saved = 0
        
        return saved, errors
    
    def _get_stats(self) -> Dict:
        """Get import statistics."""
        return {
            'imported': self.imported_count,
            'skipped': self.skipped_count,
            'errors': self.errors,
            'warnings': self.warnings
        }
    
    def get_template(self) -> Dict:
        """Get template structure for question import."""
        return {
            'required_columns': self.REQUIRED_COLUMNS,
            'optional_columns': self.OPTIONAL_COLUMNS,
            'valid_categories': self.VALID_CATEGORIES,
            'valid_levels': self.VALID_LEVELS,
            'example_row': {
                'soru_metni': 'What is the correct form of the verb?',
                'kategori': 'grammar',
                'seviye': 'B1',
                'sik_a': 'go',
                'sik_b': 'goes',
                'sik_c': 'going',
                'sik_d': 'went',
                'dogru_cevap': 'B',
                'aciklama': 'Third person singular requires -s ending',
                'tags': 'verbs, present simple'
            }
        }


def import_questions_from_file(file_path: str) -> Tuple[int, Dict]:
    """Convenience function to import questions from a file."""
    importer = QuestionImporter()
    
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext in ['.xlsx', '.xls']:
        questions, stats = importer.import_from_excel(file_path)
    elif ext == '.csv':
        questions, stats = importer.import_from_csv(file_path)
    else:
        return 0, {'errors': [f"Desteklenmeyen dosya formatı: {ext}"]}
    
    if questions:
        saved, save_errors = importer.save_questions_to_db(questions)
        stats['saved'] = saved
        stats['errors'].extend(save_errors)
    else:
        stats['saved'] = 0
    
    return stats.get('saved', 0), stats
